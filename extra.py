# -------- Documentation ----------
# python +3.11.0 install
# pip install selenium
# setup chrome.exe to system variable path in Development
# Download and place to C:\ driver,  https://googlechromelabs.github.io/chrome-for-testing/#stable 

from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import ui
from time import sleep
from threading import Thread
from openpyxl import Workbook
import json

def Find_Element(driver : webdriver.Chrome, by, value : str) -> WebElement:
    while True:
        try:
            element = driver.find_element(by, value)
            break
        except:
            pass
        sleep(0.1)
    return element

def Find_Elements(driver : webdriver.Chrome, by, value : str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(by, value)
            if len(elements) > 0:
                break
        except:
            pass
        sleep(0.1)
    return elements

def Send_Keys(element : WebElement, content : str):
    element.clear()
    for i in content:
        element.send_keys(i)
        sleep(0.1)

service = Service(executable_path="C:\chromedriver-win64\chromedriver.exe")   
options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9030")
driver = webdriver.Chrome(service=service, options=options)

wb = Workbook()
sheet = wb.active
item = ['Product Link', 'Title', 'Price', 'Item Number', 'Sub-Partnumbers', 'Size', 'Resolution', 'Panel', 'Surface', 'Frame rate', 'Backlight', 'Length/Width', 'Thickness', 'Brackets', 'Position of display connector', 'Width of display connector', 'Number of pins', 'Displayansteuerung', 'Excerpt of suitable models for P/N']
for i in range(1, 20):
    sheet.cell(row = 1, column = i).value = item[i-1]

with open('1001-2000.json', 'r') as file:
    links = json.load(file)

start_row = 2
for item_index, item in enumerate(links):
    driver.get(item["product link"])
    sheet.cell(row = start_row, column = 1).value = item["product link"]
    try:
        title = driver.find_element(By.CLASS_NAME, 'product-title').text
        sheet.cell(row = start_row, column = 2).value = title
        print(f'title : {title}')
    except:
        pass
    try:
        price = driver.find_element(By.CLASS_NAME, 'product-price').find_element(By.TAG_NAME, 'span').text.split(' ')
        sheet.cell(row = start_row, column = 3).value = price[0]
        print(f'price : {price[0]}')
    except:
        pass
    try:
        item_number = driver.find_element(By.XPATH, '//*[@id="readMoreProductInfo"]/dl[1]/dd[1]').text
        sheet.cell(row = start_row, column = 4).value = item_number
        print(f'Item number : {item_number}')
    except:
        pass
    try:
        excerpt = driver.find_element(By.XPATH, '//*[@id="readMoreProductInfo"]/dl[3]/dd').text
        sheet.cell(row = start_row, column = 19).value = excerpt
        print(f'excerpt of suitable models for P/N : {excerpt}')
    except:
        pass
    wb.save('output.xlsx')
    start_row += 1